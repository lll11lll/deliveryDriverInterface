# Nick Greiner
# 01/17/2023
# Button frame class for the driver interface, allows the user to update the total, or average table.
# to the right of that are two button that provide the latest entry and highest tip amount

import tkinter as tk
from tkinter import ttk
import openpyxl


class ButtonFrame(ttk.LabelFrame):
    def __init__(self, mainframe):
        super().__init__(mainframe)

        # Giving the frame a title
        self['text'] = 'Buttons'

        options = {'padx': 10, 'pady': 5}

        # Adding the buttons to the buttonFrame
        self.updateTotalTableButton = tk.Button(
            self, text='Update Total Table', width=30, bg='purple')

        self.updateAverageTableButton = tk.Button(
            self, text='Update Average Table', width=30, bg='purple')

        self.getLatestDateButton = tk.Button(
            self, text='Latest Entry', width=15, bg='purple')

        self.getHighestTipButton = tk.Button(
            self, text='Highest Tip', width=15, bg='purple')

        # Button functions
        # When any button is pressed a pop up appears that either provides the user info on their highest tip, latest entry, or succesful updating of a table.
        def popUp(msg, title):
            top = tk.Tk()  # Basic window

            top.title(title)
            myLabel = ttk.Label(top, text=msg, width=50)

            def buttonClick():
                top.destroy()

            B2 = tk.Button(top, text="Ok", width=20, padx=20,
                           pady=20, command=buttonClick, fg="black")

            myLabel.grid(row=0, column=0, columnspan=2, **options)
            B2.grid(row=1, column=0, columnspan=2)

        # Updates the total table, which includes: Total income, Total wage, and Total tip
        def updateTotalTable():
            workbook = openpyxl.load_workbook('myDriverData2023.xlsx')
            sheet = workbook['mainSheet']
            count = 3 + int(sheet['N3'].value)
            # I could use a dictionary here and a for loop to make this look neater

            # Total Income
            sheet['K3'].value = f'=SUM(I3:I{count})'

            # Total Wage
            sheet['L3'].value = f'=SUM(F3:F{count})'

            # Total Tip
            sheet['M3'].value = f'=SUM(B3:B{count})'

            workbook.save('myDriverData2023.xlsx')
            workbook.close()
            popUp("'The excel table 'totalTable' was sucessfully updated!", 'Sucess!')
            self.updateTotalTableButton.configure(bg='green')

        # Updates the average table, which include: Average Hourly Rate, Average Daily Tip Amount
        # Average $ I make per delivery, Average hours per shift I worked, Average # of deliveries per shift
        def updateAverageTable():
            workbook = openpyxl.load_workbook('myDriverData2023.xlsx')
            sheet = workbook['mainSheet']
            count = 2 + int(sheet['N3'].value)
            # I could use a dictionary here and a for loop to make this look neater
            # Hourly Rate
            sheet['P3'].value = f'=AVERAGE(H3:H{count})'
            # Daily Tip amount
            sheet['Q3'].value = f'=AVERAGE(B3:B{count})'
            # Per Delivery
            sheet['R3'].value = f'=AVERAGE(G3:G{count})'
            # hours
            sheet['S3'].value = f'=AVERAGE(C3:C{count})'
            # deliveries
            sheet['T3'].value = f'= AVERAGE(D4:D{count})'

            workbook.save('myDriverData2023.xlsx')
            workbook.close()
            popUp("The excel table 'averageTable' was sucessfully updated!", 'Sucess!')
            self.updateAverageTableButton.configure(bg='green')

        # provides the latest dates, loops through the start of the date column to the
        # end, adds every value to dateList, then returns the last value
        def getLatestEntry():
            workbook = openpyxl.load_workbook('myDriverData2023.xlsx')
            sheet = workbook['mainSheet']
            count = 2 + int(sheet['N3'].value)
            dateList = []
            for row in sheet.iter_rows(min_row=3, max_col=1, max_row=count):
                for cell in row:
                    dateList.append(cell.value)
            workbook.close()
            return f' The latest date entry is {dateList[-1]}'

        # Provides the highest tip, loops through the start of the tip colomn to the
        # end, adds every tip to tipList, then returns the max value of the list

        def getHighestTip():
            workbook = openpyxl.load_workbook('myDriverData2023.xlsx')
            sheet = workbook['mainSheet']
            count = 2 + int(sheet['N3'].value)
            tipList = []
            for row in sheet.iter_rows(min_row=3, min_col=2, max_col=2, max_row=count):
                for cell in row:
                    tipList.append(int(cell.value))
            return f' Your highest tip was {max(tipList)}'

        # Adding functions to buttons
        self.updateTotalTableButton.configure(command=updateTotalTable)

        self.updateAverageTableButton.configure(command=updateAverageTable)

        self.getLatestDateButton.configure(
            command=lambda: popUp(getLatestEntry(), 'Latest Entry'))

        self.getHighestTipButton.configure(
            command=lambda: popUp(getHighestTip(), 'Highest Tip'))

        # Adding it to the grid inside of the button frame
        self.updateTotalTableButton.grid(
            row=0, column=0, rowspan=2, columnspan=2, **options)
        self.updateAverageTableButton.grid(
            row=2, column=0, rowspan=2, columnspan=2, **options)
        self.getLatestDateButton.grid(row=0, column=2, **options)
        self.getHighestTipButton.grid(row=2, column=2, **options)

        # Adding button frame to the mainFrame
        self.grid(row=1, column=0, columnspan=3,
                  sticky="ew", **options)
