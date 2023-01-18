# Nick Greiner
# 01/15/2023
# Created to easily track my drivering tips and hourly rate
# Using tkinter to create a nice looking gui, and using openpyxl to
# transfer the data to the excel sheet

from tkinter import *
import sv_ttk
import openpyxl
import subprocess

# Opening the excel sheet
workbook = openpyxl.load_workbook('myDriverData2023.xlsx')
sheet = workbook['mainSheet']


def popUp(msg, title):
    top = Tk()  # Basic window
    top.title(title)
    myLabel = Label(top, text=msg, width=50, padx=10, pady=10)

    def buttonClick():
        top.destroy()

    B2 = Button(top, text="Ok", width=20, padx=20,
                pady=20, command=buttonClick, fg="black")

    myLabel.grid(row=0, column=0, columnspan=2)
    B2.grid(row=1, column=0, columnspan=2)


def openFile():
    subprocess.Popen('myDriverData2023.xlsx', shell=True)


def getLatestEntry():
    count = 2 + int(sheet['N3'].value)
    dateList = []
    for row in sheet.iter_rows(min_row=3, max_col=1, max_row=count):
        for cell in row:
            dateList.append(cell.value)
    return f' The latest date entry is {dateList[-1]}'


def getHighestTip():
    count = 2 + int(sheet['N3'].value)
    tipList = []
    for row in sheet.iter_rows(min_row=3, min_col=2, max_col=2, max_row=count):
        for cell in row:
            tipList.append(int(cell.value))
    return f' Your highest tip was {max(tipList)}'

# Creating the gui interface


def myWindow():
    # Creating my window
    root = Tk()
    root.title('Delivery Driver Interface')

    frame = Frame(root)

    dataEntryFrame = LabelFrame(frame, text='Data Entry')
    dateLabel = Label(dataEntryFrame, text='Date')
    tipLabel = Label(dataEntryFrame, text='Tip Amount')
    hourLabel = Label(dataEntryFrame, text='Hours Worked')
    deliveryLabel = Label(dataEntryFrame, text='Deliveries Taken')
    tempLabel = Label(
        dataEntryFrame, text=' ')

    dateEntry = Entry(dataEntryFrame)
    dateEntry.insert(0, 'MM/DD/YYYY')

    tipEntry = Entry(dataEntryFrame)
    hourEntry = Entry(dataEntryFrame)
    deliveryEntry = Entry(dataEntryFrame)

    addToExcel = Button(dataEntryFrame, text='Add date to excel',
                        bg='green', padx=20, width=10)
    # Function for when the button is pressed

    def addDataToExcel():
        count = 3 + int((sheet['N3'].value))
        tip = int(tipEntry.get())
        hour = float(hourEntry.get())
        delivery = int(deliveryEntry.get())

        myEntries = 1
        myDataList = []
        myDataList.append(dateEntry.get())
        myDataList.append(tip)
        myDataList.append(hour)
        myDataList.append(delivery)

        statList = []

        wageAmount = round((hour * 7.98), 2)
        tipPerDelivery = round((tip / delivery), 2)
        hourlyRate = round((wageAmount + tip) / hour, 2)
        dailyTotal = round((wageAmount + tip), 2)

        statList.append(wageAmount)
        statList.append(tipPerDelivery)
        statList.append(hourlyRate)
        statList.append(dailyTotal)

        dateEntry.delete(0, END)
        dateEntry.insert(0, 'MM/DD/YYYY')
        tipEntry.delete(0, END)
        hourEntry.delete(0, END)
        deliveryEntry.delete(0, END)

        while myEntries != 5:
            sheet.cell(
                row=count, column=myEntries).value = myDataList[myEntries - 1]
            sheet.cell(
                row=count, column=myEntries + 5).value = statList[myEntries - 1]
            myEntries += 1

        sheet['N3'].value += 1
        # myDataList.clear()
        sheet.tables['dataTable'].ref = f'A2:D{count}'
        sheet.tables['statisticTable'].ref = f'F2:I{count}'
        workbook.save('myDriverData2023.xlsx')

    addToExcel.configure(command=addDataToExcel)

    buttonFrame = LabelFrame(frame, text='Buttons')

    updateTotalTableButton = Button(
        buttonFrame, text='Update Total Table', bg='purple', padx=20, width=30)

    def updateTotalTable():
        count = 3 + int(sheet['N3'].value)

        # I could use a dictionary here and a for loop to make this look neater
        # Total Income

        sheet['K3'].value = f'=SUM(I3:I{count})'

        # Total Wage
        sheet['L3'].value = f'=SUM(F3:F{count})'

        # Total Tip
        sheet['M3'].value = f'=SUM(B3:B{count})'

        workbook.save('myDriverData2023.xlsx')
        popUp("'The excel table 'totalTable' was sucessfully updated!", 'Sucess!')
        updateTotalTableButton.configure(bg='green')

    updateAverageTableButton = Button(
        buttonFrame, text='Update Average Table', bg='purple', padx=20, width=30)

    def updateAverageTable():
        count = 2 + int(sheet['N3'].value)
        # I could use a dictionary here and a for loop to make this look neater
        # Hourly Rate
        sheet['P3'].value = f'=AVERAGE(H3:H{count})'
        # Daily Tip amount
        sheet['Q3'].value = f'=AVERAGE(B3:B{count})'
        # Per Delivery
        sheet['R3'].value = f'=AVERAGE(G3:G{count})'
        # hours
        sheet['S3'].value = f'=AVERAGE(C3:C{count})'
        # deliveries
        sheet['T3'].value = f'= AVERAGE(D4:D{count})'

        workbook.save('myDriverData2023.xlsx')
        popUp("The excel table 'averageTable' was sucessfully updated!", 'Sucess!')
        updateAverageTableButton.configure(bg='green')

    getLatestDateButton = Button(
        buttonFrame, text='Latest entry', bg='purple', width=15, padx=20, command=lambda: popUp(getLatestEntry(), 'Latest entry'))
    getHighestTipButton = Button(
        buttonFrame, text='Highest Tip', bg='purple', width=15, padx=20, command=lambda: popUp(getHighestTip(), 'Highest tip'))

    updateTotalTableButton.configure(command=updateTotalTable)
    updateAverageTableButton.configure(command=updateAverageTable)

    finalFrame = LabelFrame(frame, text='Open Excel / Quit')
    openExcelSheetButton = Button(
        finalFrame, text="Open Excel", command=openFile, width=50, padx=20, bg='blue')
    quitButton = Button(finalFrame, text="Quit", width=50, padx=20,
                        bg='red', command=root.destroy)

    buttonFrame.grid(row=1, column=0, columnspan=3)
    dataEntryFrame.grid(row=0, column=0, columnspan=3)
    finalFrame.grid(row=2, column=0, columnspan=3)

    # Grid within the dataEntryFrame, in dataEntryFrame
    dateLabel.grid(row=1, column=0)
    dateEntry.grid(row=2, column=0)

    tipLabel.grid(row=1, column=1)
    tipEntry.grid(row=2, column=1)

    tempLabel.grid(row=3, column=0, columnspan=2)
    hourLabel.grid(row=4, column=1)
    hourEntry.grid(row=5, column=1)

    deliveryLabel.grid(row=4, column=0)
    deliveryEntry.grid(row=5, column=0)

    addToExcel.grid(row=3, column=3, rowspan=1, columnspan=1)

    # Grid within the dataEntryFrame, in button frame
    updateTotalTableButton.grid(row=1, column=0, columnspan=2)
    updateAverageTableButton.grid(row=2, column=0, columnspan=2)
    getLatestDateButton.grid(row=1, column=2,)
    getHighestTipButton.grid(row=2, column=2)

    # Grid within the dataEntryFrame, in final frame
    openExcelSheetButton.grid(row=0, column=0, rowspan=2, columnspan=2)
    quitButton.grid(row=2, column=0, rowspan=2, columnspan=2)
    frame.pack()

    for widget in dataEntryFrame.winfo_children():
        widget.grid_configure(padx=10, pady=5)

    for widget in buttonFrame.winfo_children():
        widget.grid_configure(padx=10, pady=5)

    for widget in finalFrame.winfo_children():
        widget.grid_configure(padx=10, pady=5)
    sv_ttk.set_theme('dark')
    root.mainloop()


myWindow()
