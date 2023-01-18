# Nick Greiner
# 01/16/2023
# Data Entry frame class for the driver interface, contains the labels and entries
# for the date, tip amount, deliveries taken and hours worked.
# finally it has a "add date to excel" button.

import tkinter as tk
from tkinter import ttk
import openpyxl


class DataEntryFrame(ttk.LabelFrame):
    def __init__(self, mainFrame):
        super().__init__(mainFrame)
        self['text'] = 'Data Entry'

        options = {'padx': 10, 'pady': 5}
        # Label for date, tip, hour, and delivery amount

        self.dateLabel = ttk.Label(self, text='Date')
        self.dateEntry = ttk.Entry(self)
        self.dateEntry.insert(0, 'MM/DD/YYYY')

        self.tip = tk.IntVar()
        self.tipLabel = ttk.Label(self, text='Tip')
        self.tipEntry = ttk.Entry(self, textvariable=self.tip)

        self.hour = tk.DoubleVar()
        self.hourLabel = ttk.Label(self, text='Hours')
        self.hourEntry = ttk.Entry(self, textvariable=self.hour)

        self.deliveries = tk.IntVar()
        self.deliveryLabel = ttk.Label(
            self, text='Deliveries Taken')
        self.deliveryEntry = ttk.Entry(self, textvariable=self.deliveries)

        # Button and command for adding to excel
        self.addToExcelButton = tk.Button(
            self, text='Add to Excel', bg='green')

        # adding to excel function
        def addToExcel():
            workbook = openpyxl.load_workbook('myDriverData2023.xlsx')
            sheet = workbook['mainSheet']
            count = 3 + int((sheet['N3'].value))

            tip = int(self.tipEntry.get())
            hour = float(self.hourEntry.get())
            delivery = int(self.deliveryEntry.get())

            myEntries = 1
            myDataList = []
            myDataList.append(self.dateEntry.get())
            myDataList.append(tip)
            myDataList.append(hour)
            myDataList.append(delivery)

            statList = []

            wageAmount = round((hour * 7.98), 2)
            tipPerDelivery = round((tip / delivery), 2)
            hourlyRate = round((wageAmount + tip) / hour, 2)
            dailyTotal = round((wageAmount + tip), 2)

            statList.append(wageAmount)
            statList.append(tipPerDelivery)
            statList.append(hourlyRate)
            statList.append(dailyTotal)

            self.dateEntry.delete(0, 'end')
            self.dateEntry.insert(0, 'MM/DD/YYYY')
            self.tipEntry.delete(0, 'end')
            self.hourEntry.delete(0, 'end')
            self.deliveryEntry.delete(0, 'end')

            while myEntries != 5:
                sheet.cell(
                    row=count, column=myEntries).value = myDataList[myEntries - 1]
                sheet.cell(
                    row=count, column=myEntries + 5).value = statList[myEntries - 1]
                myEntries += 1

            sheet['N3'].value += 1
            sheet.tables['dataTable'].ref = f'A2:D{count}'
            sheet.tables['statisticTable'].ref = f'F2:I{count}'
            workbook.save('myDriverData2023.xlsx')
            workbook.close()

        # Adding the command to the button so when it is pressed it adds the data that the user input and stores it into the myDriverData excel sheet
        self.addToExcelButton.configure(command=addToExcel)

        # Add labels, entries, and button to the dataEntry Label frame
        self.dateLabel.grid(row=0, column=0, **options)
        self.dateEntry.grid(row=1, column=0, **options)
        self.tipLabel.grid(row=0, column=1, **options)
        self.tipEntry.grid(row=1, column=1, **options)
        self.hourLabel.grid(row=3, column=0, **options)
        self.hourEntry.grid(row=4, column=0, **options)
        self.deliveryLabel.grid(row=3, column=1, **options)
        self.deliveryEntry.grid(row=4, column=1, **options)
        self.addToExcelButton.grid(row=2, column=3, rowspan=2, **options)
        self.grid(row=0, column=0, columnspan=3, sticky="ew", **options)
